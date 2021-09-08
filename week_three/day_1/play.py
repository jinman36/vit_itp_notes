import openpyxl

#Assign your .xlsx file to a variable
filename = "./week_three/day_1/lecture.xlsx"
wb = openpyxl.load_workbook(filename)

print(type(wb))

sheet_name = wb.sheetnames
# print(sheet_name)

my_sheet = wb['Sheet1']
# wb.get_sheet_by_name(sheet_name[0]) - another way to access the sheet by index
# print(my_sheet)

# apple_cell = my_sheet['B1']
# print(apple_cell.value)


# strawberry_cell = my_sheet['B']
# print(strawberry_cell.value)

# print('Row %s, Column %s is %s' %(apple_cell.row, apple_cell.column, apple_cell.value))

# for row in my_sheet(A1,C7):
#   print(row.value)

for i in range(1, my_sheet.max_row + 1):
    # print(x, my_sheet.cell(row = x, column = 1).value)
    date = "A" + str(i)
    date_cell = my_sheet[date]

    amount = "C" + str(i)
    amount_cell = my_sheet[amount]

    fruit = "B" + str(i)
    fruit_cell = my_sheet[fruit]

    print ("On the date of " + str(date_cell.value.date()) + ", " + str(amount_cell.value) + " amount of " + str(fruit_cell.value) + "!")
# for i in range(1, 8):
#     print(i, my_sheet.cell(row = i, column =2).value)
# for y in range(1, 8):
#     print(y, my_sheet.cell(row = y, column =3).value)

# print('On the date')