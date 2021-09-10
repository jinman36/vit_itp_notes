import os
import openpyxl
from openpyxl.workbook.workbook import Workbook

wb = openpyxl.load_workbook('week_three/day_2/lecture.xlsx')
# print(wb.workbook)

sheet_name = wb['Population by Census Tract']

# print(sheet_name)

working_sheet = sheet_name

print(working_sheet)

# for i in range(1, working_sheet.max_row+1):
#   pop = "D" + str(i)
#   population = working_sheet[pop]
#   print(population.value)


  
# wb = Workbook()
# print(str(wb))

# for i in range(3, 10):
#   wb.create_sheet()

# for sheet in wb:
#   y = 5
#   sheet.title = 'MyNewSheet' + str(y)
#   y += 1

# for i in range (1, wb)
